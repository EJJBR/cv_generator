import argparse
from pathlib import Path

from openpyxl import Workbook, load_workbook


HEADER_ALIASES_FOTO = (
    "foto de perfil (sólo jpg)",
    "foto",
    "foto drive",
    "foto de perfil",
    "fotos",
)


def normalizar_header(valor):
    if valor is None:
        return ""
    return " ".join(str(valor).strip().lower().split())


def encontrar_columna_foto(encabezados):
    for idx, encabezado in enumerate(encabezados):
        nombre = normalizar_header(encabezado)
        if nombre in HEADER_ALIASES_FOTO:
            return idx
        if "foto" in nombre and ("perfil" in nombre or "drive" in nombre or "upload" in nombre):
            return idx
    return None


def fila_vacia(fila):
    return not any((valor is not None and str(valor).strip() != "") for valor in fila)


def limpiar_excel(ruta_entrada: str, ruta_salida: str | None = None):
    ruta_entrada = Path(ruta_entrada)
    if not ruta_entrada.exists():
        raise FileNotFoundError(f"No existe el archivo: {ruta_entrada}")

    wb = load_workbook(ruta_entrada, data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        raise ValueError(f"El archivo {ruta_entrada} está vacío.")

    encabezados = list(filas[0])
    registros = []
    for fila in filas[1:]:
        if fila_vacia(fila):
            continue
        registros.append(fila)

    if not registros:
        raise ValueError("No se encontraron registros útiles después de quitar filas vacías.")

    # Asegurarnos de que el largo del registro coincida con el número de columnas
    ancho = len(encabezados)
    for i, fila in enumerate(registros):
        registro = list(fila)
        if len(registro) < ancho:
            registro += [None] * (ancho - len(registro))
        registros[i] = registro

    # ID único y columna de fotos
    nueva_wb = Workbook()
    nueva_ws = nueva_wb.active
    nueva_ws.append(["ID"] + [str(h).strip() if h is not None else "" for h in encabezados])

    faltantes = []
    indice_foto = encontrar_columna_foto(nueva_ws[1])
    if indice_foto is None:
        indice_foto = encontrar_columna_foto(encabezados)

    for idx, fila in enumerate(registros, start=1):
        fila_con_id = [idx] + list(fila)
        nueva_ws.append(fila_con_id)

        if indice_foto is None:
            continue

        valor_foto = fila[indice_foto] if indice_foto < len(fila) else None
        if valor_foto is None or str(valor_foto).strip() == "":
            nombre = fila[1] if len(fila) > 1 else ""
            faltantes.append({"id": idx, "nombre": nombre})

    if ruta_salida is None:
        ruta_salida = ruta_entrada.with_name(f"{ruta_entrada.stem}_limpio.xlsx")
    else:
        ruta_salida = Path(ruta_salida)

    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    nueva_wb.save(ruta_salida)

    print(f"✅ Archivo limpio generado: {ruta_salida}")
    print(f"📊 Registros totales: {len(registros)}")
    print(f"⚠️ Registros sin link de imagen: {len(faltantes)}")

    if faltantes:
        print("\nRegistros faltantes:")
        for item in faltantes[:10]:
            print(f"- ID {item['id']}: {item['nombre']}")
        if len(faltantes) > 10:
            print(f"... y {len(faltantes) - 10} más")

    return {
        "ruta_salida": str(ruta_salida),
        "total": len(registros),
        "faltantes": faltantes,
    }


def main():
    parser = argparse.ArgumentParser(description="Limpia un Excel de formulario, elimina filas vacías y agrega IDs.")
    parser.add_argument("ruta_entrada", nargs="?", default=r"C:\Users\LENOVO\Downloads\Hoja de Vida Docente (respuestas) (2).xlsx")
    parser.add_argument("--salida", default=None)
    args = parser.parse_args()

    limpiar_excel(args.ruta_entrada, args.salida)


if __name__ == "__main__":
    main()
