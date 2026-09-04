"""
controllers/masivo_controller.py
Lógica para procesamiento masivo de CVs.
"""

import os
import re
import threading
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

import openpyxl
import requests

from data_reader import leer_excel
from pdf_generator import generar_cv
from ui.config import OUTPUT_DIR, OUTPUT_IMAGENES, OUTPUT_LOGS, OUTPUT_REGISTROS, LOGO_PATH
from ui.utils import nombre_archivo_pdf


def _slug_nombre(nombre: str) -> str:
    if not nombre:
        return "sin_nombre"
    nombre = str(nombre)
    nombre = re.sub(r"[^a-zA-Z0-9]+", "_", nombre).strip("_")
    return nombre or "sin_nombre"


def _nombre_archivo_imagen(nombre: str) -> str:
    """Conserva el nombre legible y elimina caracteres inválidos en Windows."""
    nombre = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "", str(nombre or ""))
    nombre = " ".join(nombre.split())
    return (nombre.strip(" .") or "SIN_NOMBRE").upper()


def _max_workers() -> int:
    """Limita la concurrencia a cuatro tareas y se adapta al equipo disponible."""
    return max(1, min(4, os.cpu_count() or 1))


def _generar_cv_en_proceso(datos: dict, ruta_pdf: str):
    """Genera un CV aislado para que pueda ejecutarse en otro proceso."""
    generar_cv(datos, ruta_pdf, logo_path=LOGO_PATH)


class MasivoController:
    """Controlador para generación masiva de CVs."""
    
    def __init__(self, callback_log, callback_agregar_pendiente, callback_completar):
        """
        callback_log: función para escribir en el log (texto)
        callback_agregar_pendiente: función para agregar docente sin foto (dict)
        callback_completar: función al terminar el proceso
        """
        self.callback_log = callback_log
        self.callback_agregar_pendiente = callback_agregar_pendiente
        self.callback_completar = callback_completar

    def transformar_excel(self, ruta_excel: str) -> tuple[bool, str, str]:
        """Limpia el Excel, elimina filas vacías y agrega la columna ID."""
        try:
            ruta_excel = str(ruta_excel)
            if not os.path.exists(ruta_excel):
                return False, "❌ No existe el archivo Excel seleccionado.", ""

            os.makedirs(OUTPUT_REGISTROS, exist_ok=True)
            wb = openpyxl.load_workbook(ruta_excel, data_only=True)
            ws = wb.active
            filas = list(ws.iter_rows(values_only=True))
            if not filas:
                return False, "❌ El Excel está vacío.", ""

            encabezados = [str(h).strip() if h is not None else "" for h in filas[0]]
            registros = []
            for fila in filas[1:]:
                if not fila or not any((valor is not None and str(valor).strip() != "") for valor in fila):
                    continue
                registros.append(list(fila))

            if not registros:
                return False, "❌ No se encontraron registros útiles después de limpiar filas vacías.", ""

            nombre_base = os.path.splitext(os.path.basename(ruta_excel))[0]
            output_path = os.path.join(OUTPUT_REGISTROS, f"{nombre_base}_limpio.xlsx")

            nueva_wb = openpyxl.Workbook()
            nueva_ws = nueva_wb.active
            nueva_ws.append(["ID"] + encabezados)

            for idx, fila in enumerate(registros, start=1):
                if len(fila) < len(encabezados):
                    fila = list(fila) + [None] * (len(encabezados) - len(fila))
                nueva_ws.append([idx] + fila)

            nueva_wb.save(output_path)
            return True, f"✅ Excel transformado en {output_path}", output_path
        except Exception as exc:
            return False, f"❌ Error al transformar Excel: {exc}", ""

    def descargar_fotos_excel(self, ruta_excel: str) -> tuple[bool, str]:
        """Descarga imágenes usando hasta cuatro hilos para tareas de red."""
        ok, resumen, _ = self.descargar_fotos_excel_detallado(ruta_excel)
        return ok, resumen

    def descargar_fotos_excel_detallado(self, ruta_excel: str) -> tuple[bool, str, list[dict]]:
        """Descarga imágenes y retorna el estado por cada fila para la interfaz."""
        try:
            os.makedirs(OUTPUT_IMAGENES, exist_ok=True)
            os.makedirs(OUTPUT_LOGS, exist_ok=True)

            wb = openpyxl.load_workbook(ruta_excel, data_only=True)
            filas = list(wb.active.iter_rows(values_only=True))
            if not filas:
                return False, "❌ El Excel transformado está vacío.", []

            encabezados = [str(h).strip() if h is not None else "" for h in filas[0]]
            encabezados_norm = [re.sub(r"[^a-z0-9]+", " ", str(h).lower().strip()) for h in encabezados]

            idx_id = next((i for i, nombre in enumerate(encabezados_norm)
                           if nombre in {"id", "identificacion", "numero de identificacion"}
                           or nombre.startswith("id ")), None)
            idx_nombre = next(
                (i for i, nombre in enumerate(encabezados_norm)
                 if ("apellidos" in nombre and "nombre" in nombre)
                 or nombre in {"nombre completo", "apellidos", "nombres", "nombre"}),
                None,
            )
            idx_foto = next(
                (i for i, nombre in enumerate(encabezados_norm)
                 if "foto" in nombre and ("perfil" in nombre or "drive" in nombre or "jpg" in nombre or "link" in nombre)),
                None,
            )
            if idx_id is None or idx_nombre is None or idx_foto is None:
                return False, "❌ No se encontraron las columnas ID, Apellidos y Nombres o Foto de Perfil.", []

            log_path = os.path.join(OUTPUT_LOGS, "descarga_fotos.log")
            tareas = []
            total = 0
            faltantes = 0
            errores = 0
            detalle = []

            for fila in filas[1:]:
                if not fila or not any(valor is not None and str(valor).strip() != "" for valor in fila):
                    continue
                id_val = fila[idx_id] if idx_id < len(fila) else None
                if id_val is None or str(id_val).strip() == "":
                    continue

                total += 1
                foto_val = fila[idx_foto] if idx_foto < len(fila) else None
                url = str(foto_val).strip() if foto_val is not None else ""
                nombre_base = fila[idx_nombre] if idx_nombre < len(fila) else ""
                if isinstance(nombre_base, str):
                    nombre = nombre_base
                else:
                    nombre = str(nombre_base or "")

                if encabezados_norm[idx_nombre] in {"apellidos", "nombres"}:
                    nombre_comp = []
                    for j, col in enumerate(encabezados_norm):
                        if j == idx_nombre:
                            continue
                        if col in {"apellidos", "nombres", "nombre completo", "apellidos y nombres", "nombre"}:
                            valor = fila[j] if j < len(fila) else ""
                            if valor is not None and str(valor).strip() != "":
                                nombre_comp.append(str(valor).strip())
                    if nombre_comp:
                        nombre = " ".join(nombre_comp)

                item = {
                    "id": str(id_val).strip(),
                    "nombre": nombre.strip() or "Sin nombre",
                    "tiene_enlace": bool(url),
                    "descargada": False,
                    "fallida": False,
                    "can_retry": False,
                    "progreso": 0,
                    "estado": "Sin enlace",
                    "ruta_foto": "",
                }

                if not url or not any(dominio in url.lower() for dominio in (
                    "drive.google.com", "docs.google.com", "googleusercontent"
                )):
                    faltantes += 1
                    item["estado"] = "Sin enlace"
                    item["tiene_enlace"] = False
                    item["progreso"] = 0
                    detalle.append(item)
                    self._registrar_descarga(log_path, f"⚠️ ID {id_val}: falta link de Drive en la columna de foto.")
                    continue

                match = re.search(r"/d/([a-zA-Z0-9_-]+)", url) or re.search(r"[?&]id=([a-zA-Z0-9_-]+)", url)
                if not match:
                    errores += 1
                    item["estado"] = "Error al extraer enlace"
                    item["fallida"] = True
                    item["can_retry"] = True
                    detalle.append(item)
                    self._registrar_descarga(log_path, f"❌ ID {id_val}: no se pudo extraer el file_id del link de Drive.")
                    continue

                item["estado"] = "Descargando"
                item["progreso"] = 40
                item["tiene_enlace"] = True
                nombre_archivo = _nombre_archivo_imagen(nombre)
                destino = os.path.join(OUTPUT_IMAGENES, f"{id_val}_{nombre_archivo}.jpg")
                item["ruta_foto"] = destino
                tareas.append((str(id_val), match.group(1), destino, item))
                detalle.append(item)

            with ThreadPoolExecutor(max_workers=_max_workers()) as executor:
                futuros = {
                    executor.submit(self._descargar_foto, file_id, destino): (id_val, destino, item)
                    for id_val, file_id, destino, item in tareas
                }
                for futuro in as_completed(futuros):
                    id_val, destino, item = futuros[futuro]
                    try:
                        futuro.result()
                        item["estado"] = "Descargada"
                        item["descargada"] = True
                        item["tiene_enlace"] = True
                        item["progreso"] = 100
                        item["fallida"] = False
                        item["can_retry"] = False
                        self._registrar_descarga(
                            log_path,
                            f"✅ ID {id_val}: descargada -> {os.path.basename(destino)}",
                        )
                    except Exception as exc:
                        errores += 1
                        item["estado"] = "Error al descargar"
                        item["fallida"] = True
                        item["can_retry"] = True
                        item["progreso"] = 0
                        self._registrar_descarga(log_path, f"❌ ID {id_val}: error al descargar la foto -> {exc}")

            resumen = (
                f"\n==== RESUMEN ====\n"
                f"Registros con ID: {total}\n"
                f"Descargadas: {sum(1 for fila in detalle if fila.get('descargada'))}\n"
                f"Faltantes de link: {faltantes}\n"
                f"Errores de descarga: {errores}\n"
                f"Log: {log_path}\n"
            )
            self._registrar_descarga(log_path, resumen.rstrip())
            return True, resumen, detalle
        except Exception as exc:
            return False, f"❌ Error intentando descargar imágenes: {exc}", []

    def _descargar_foto(self, file_id: str, destino: str):
        """Descarga una imagen individual; se ejecuta dentro de un hilo."""
        api_url = "https://docs.google.com/uc?export=download"
        session = requests.Session()
        response = session.get(api_url, params={"id": file_id}, stream=True, timeout=60)
        token = next((value for key, value in response.cookies.items() if key.startswith("download_warning")), None)
        if token:
            response = session.get(api_url, params={"id": file_id, "confirm": token}, stream=True, timeout=60)
        if response.status_code != 200:
            raise RuntimeError(f"Google Drive respondió HTTP {response.status_code}")
        if "image" not in response.headers.get("Content-Type", "").lower():
            raise RuntimeError(f"La URL no devolvió una imagen válida: {response.headers.get('Content-Type')}")
        with open(destino, "wb") as archivo:
            for chunk in response.iter_content(32768):
                if chunk:
                    archivo.write(chunk)

    def _registrar_descarga(self, log_path: str, mensaje: str):
        """Escribe en archivo, consola y log visible de la interfaz."""
        print(mensaje)
        self.callback_log(mensaje)
        with open(log_path, "a", encoding="utf-8") as archivo:
            archivo.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {mensaje}\n")

    def procesar(self, ruta_excel: str, ruta_fotos: str):
        """Inicia procesamiento en thread separado."""
        threading.Thread(
            target=self._procesar_interno,
            args=(ruta_excel, ruta_fotos),
            daemon=True
        ).start()

    def _procesar_interno(self, ruta_excel: str, ruta_fotos: str):
        """Lógica interna de procesamiento."""
        try:
            completos, sin_foto = leer_excel(ruta_excel, ruta_fotos)
        except Exception as e:
            self.callback_log(f"❌ Error leyendo Excel: {e}")
            self.callback_completar()
            return

        total = len(completos) + len(sin_foto)
        self.callback_log(
            f"📊 {total} registros encontrados — {len(completos)} con foto, "
            f"{len(sin_foto)} sin foto\n"
        )

        os.makedirs(OUTPUT_DIR, exist_ok=True)

        # Generar CVs de los que sí tienen foto en procesos separados.
        max_workers = _max_workers()
        self.callback_log(
            f"⚙ Generando {len(completos)} CVs con hasta {max_workers} procesos..."
        )
        with ProcessPoolExecutor(max_workers=max_workers) as executor:
            futuros = {}
            for datos in completos:
                nombre = datos.get("nombre", "sin_nombre")
                ruta_pdf = os.path.join(OUTPUT_DIR, nombre_archivo_pdf(nombre, datos.get("id")))
                futuro = executor.submit(_generar_cv_en_proceso, datos, ruta_pdf)
                futuros[futuro] = nombre

            for futuro in as_completed(futuros):
                nombre = futuros[futuro]
                try:
                    futuro.result()
                    self.callback_log(f"✅ {nombre}")
                except Exception as e:
                    self.callback_log(f"❌ {nombre} → {e}")

        # Agregar pendientes
        if sin_foto:
            self.callback_log(
                f"\n⚠️  {len(sin_foto)} docente(s) sin foto — selecciona manualmente:"
            )
            for datos in sin_foto:
                self.callback_agregar_pendiente(datos)

        self.callback_log("\n✔ Proceso completado.")
        self.callback_completar()

    def generar_cv_manual(self, datos: dict) -> tuple[bool, str]:
        """
        Genera CV para docente sin foto (seleccionada manualmente).
        Retorna: (éxito: bool, mensaje: str)
        """
        nombre = datos.get("nombre", "sin_nombre")
        try:
            os.makedirs(OUTPUT_DIR, exist_ok=True)
            ruta_pdf = os.path.join(OUTPUT_DIR, nombre_archivo_pdf(nombre, datos.get("id")))
            generar_cv(datos, ruta_pdf, logo_path=LOGO_PATH)
            return True, f"✅ ID {datos.get('id', 'Sin ID')}: {nombre} → generado manualmente"
        except Exception as e:
            return False, f"❌ ID {datos.get('id', 'Sin ID')}: {nombre} → {e}"
