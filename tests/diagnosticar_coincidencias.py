import argparse
import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from src.data_reader import _buscar_foto


def _normalizar_columna(nombre: object) -> str:
    return str(nombre or "").strip().lower()


def cargar_nombres_excel(ruta_excel: str) -> list[str]:
    """Lee los nombres del Excel desde la columna 'Apellidos y Nombres'."""
    wb = load_workbook(ruta_excel, data_only=True)
    ws = wb.active

    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return []

    encabezados = filas[0]
    headers_norm = [_normalizar_columna(h) for h in encabezados]

    indice = None
    for idx, h in enumerate(headers_norm):
        if h in {"apellidos y nombres", "nombre completo", "nombres y apellidos"}:
            indice = idx
            break

    if indice is None:
        raise ValueError("No se encontró la columna 'Apellidos y Nombres' en el Excel.")

    nombres = []
    for fila in filas[1:]:
        if not fila:
            continue
        valor = fila[indice] if indice < len(fila) else None
        nombre = str(valor).strip() if valor is not None and str(valor).strip() else ""
        if nombre:
            nombres.append(nombre)

    return nombres


def listar_archivos_fotos(carpeta: str) -> list[str]:
    if not os.path.isdir(carpeta):
        raise FileNotFoundError(f"No existe la carpeta de fotos: {carpeta}")

    archivos = []
    for nombre in sorted(os.listdir(carpeta)):
        ruta = os.path.join(carpeta, nombre)
        if os.path.isfile(ruta):
            archivos.append(nombre)
    return archivos


def diagnosticar_excel_fotos(ruta_excel: str, carpeta_fotos: str):
    nombres = cargar_nombres_excel(ruta_excel)
    fotos = listar_archivos_fotos(carpeta_fotos)

    usados = set()
    encontrados = []
    no_encontrados = []

    print(f"Excel: {ruta_excel}")
    print(f"Carpeta fotos: {carpeta_fotos}")
    print(f"Total registros Excel: {len(nombres)}")
    print(f"Total archivos de fotos: {len(fotos)}")
    print("-" * 100)

    print("REGISTROS CON FOTO ENCONTRADA")
    for nombre in nombres:
        ruta = _buscar_foto(nombre, carpeta_fotos)
        if ruta:
            base = os.path.basename(ruta)
            usados.add(base)
            encontrados.append((nombre, base))
            print(f"OK   | {nombre} -> {base}")

    print("-" * 100)
    print("REGISTROS SIN FOTO")
    for nombre in nombres:
        ruta = _buscar_foto(nombre, carpeta_fotos)
        if not ruta:
            no_encontrados.append(nombre)
            print(f"FAIL | {nombre}")

    print("-" * 100)
    print("IMAGENES NO USADAS")
    sin_usar = [f for f in fotos if f not in usados]
    if not sin_usar:
        print("Ninguna imagen quedó sin usar.")
    else:
        for nombre in sin_usar:
            print(f"UNUSED | {nombre}")

    print("-" * 100)
    print(f"Resumen: {len(encontrados)} con foto / {len(no_encontrados)} sin foto / {len(sin_usar)} fotos sin usar")

    return {
        "encontrados": encontrados,
        "no_encontrados": no_encontrados,
        "sin_usar": sin_usar,
    }


def main():
    parser = argparse.ArgumentParser(
        description="Diagnóstico de coincidencias entre Excel y carpeta de fotos."
    )
    parser.add_argument("--excel", required=True, help="Ruta del archivo Excel con la columna Apellidos y Nombres.")
    parser.add_argument("--fotos", required=True, help="Ruta de la carpeta de fotos.")
    args = parser.parse_args()

    try:
        diagnosticar_excel_fotos(args.excel, args.fotos)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
